from typing import Optional, Dict, Any
from .wall_object import WallObject
from .structures import Position
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image
from gallery.models import db
from gallery.models.user import User
from .base import db

class Artwork(WallObject):
    __tablename__ = 'artworks'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    width = db.Column(db.Float, nullable=False)
    height = db.Column(db.Float, nullable=False)
    hanging_point = db.Column(db.Float, nullable=False)
    medium = db.Column(db.String(100))
    depth = db.Column(db.Float)
    price = db.Column(db.Float)
    nfs = db.Column(db.Boolean, default=False)
    notes = db.Column(db.String(500))
    image_path = db.Column(db.String(200))
    gallery_id = db.Column(db.Integer, db.ForeignKey('galleries.id'))
    x_position = db.Column(db.Float, nullable=True)
    y_position = db.Column(db.Float, nullable=True)
    wall_id = db.Column(db.Integer, db.ForeignKey('wall.id'), nullable=True)

    # New user association
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    user = db.relationship('User', back_populates='artworks')

    def __init__(self, name: str = "", medium: str = "", width: float = 0.0, height: float = 0.0,
                 depth: float = 0.0, hanging_point: float = 0.0, price: float = 0.0,
                 nfs: bool = False, image_path: str = "", notes: str = "", wall_id: Optional[int] = None,
                 x: float = 0.0, y: float = 0.0, user_id: Optional[int] = None):
        self.name = name
        self.medium = medium
        self.width = width
        self.height = height
        self.depth = depth
        self.hanging_point = hanging_point
        self.price = price
        self.nfs = nfs
        self.image_path = image_path
        self.notes = notes
        self.wall_id = wall_id
        self.x_position = x
        self.y_position = y
        self.user_id = user_id

    @property
    def position(self) -> Position:
        return Position(self.x_position, self.y_position)
    
    @position.setter
    def position(self, value: Position):
        if not isinstance(value, Position):
            raise ValueError("Position must be a Position object")
        self.x_position = value.x
        self.y_position = value.y

    def to_dict(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'name': self.name,
            'width': self.width,
            'height': self.height,
            'x_position': self.x_position,
            'y_position': self.y_position,
            'wall_id': self.wall_id,
            'medium': self.medium,
            'depth': self.depth,
            'hanging_point': self.hanging_point,
            'price': self.price,
            'nfs': self.nfs,
            'image_path': self.image_path,
            'notes': self.notes,
            'user_id': self.user_id
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Artwork':
        return cls(
            name=data.get('name', ''),
            medium=data.get('medium', ''),
            width=float(data.get('width', 0)),
            height=float(data.get('height', 0)),
            depth=float(data.get('depth', 0)),
            hanging_point=float(data.get('hanging_point', 0)),
            price=float(data.get('price', 0)),
            nfs=bool(data.get('nfs', False)),
            image_path=data.get('image_path', ''),
            notes=data.get('notes', ''),
            wall_id=data.get('wall_id'),
            x=float(data.get('position', {}).get('x', 0)),
            y=float(data.get('position', {}).get('y', 0)),
            user_id=data.get('user_id')
        )

    def export_to_excel(self, filename="artwork_export.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Artworks"
        
        headers = ["Name", "Medium", "Height", "Width", "Depth", "Hanging Point", 
                  "Value", "NFS", "Image", "Notes"]
        colors = ["FFC7CE", "FFF2CC", "FFC7CE", "FFC7CE", "FFF2CC", 
                 "FFC7CE", "FFF2CC", "FFF2CC", "FFF2CC", "FFF2CC"]
        
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        ws.append([
            self.name, self.medium, self.height, self.width, self.depth,
            self.hanging_point, self.price, self.nfs, self.image_path, self.notes
        ])
        
        if self.image_path and os.path.exists(self.image_path):
            img = Image(self.image_path)
            img.width, img.height = 100, 100
            ws.add_image(img, "I2")
        
        wb.save(filename)
        return filename

    @classmethod
    def import_from_excel(cls, filename="artwork_export.xlsx", sheetname="Artworks"):
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found: {filename}")
            
        ws = wb[sheetname]
        artworks = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, medium, height, width, depth, hanging_point, price, nfs, image_path, notes = row
            artwork = cls(
                name=name,
                medium=medium,
                height=float(height),
                width=float(width),
                depth=float(depth),
                hanging_point=float(hanging_point),
                price=float(price),
                nfs=bool(nfs),
                image_path=image_path,
                notes=notes
            )
            artworks.append(artwork)
        
        return artworks

    def __repr__(self):
        return (f"<Artwork(name='{self.name}', medium='{self.medium}', "
                f"size={self.width}x{self.height}, wall_id={self.wall_id})>")