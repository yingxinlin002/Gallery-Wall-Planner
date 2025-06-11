import openpyxl
from openpyxl.styles import Font, PatternFill
from typing import List, Optional
from .base import db
from .artwork import Artwork
from datetime import datetime

class Exhibit(db.Model):
    __tablename__ = 'exhibits'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128))
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    user = db.relationship('User', back_populates='exhibits', lazy=True)  # Changed from 'galleries'
    guest_id = db.Column(db.String(36), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    walls = db.relationship('Wall', backref='exhibit', lazy=True, cascade='all, delete-orphan')  # Changed from 'gallery'
    unplaced_artworks = db.relationship(
        'Artwork',
        primaryjoin='and_(Artwork.wall_id==None, Artwork.exhibit_id==Exhibit.id)',  # Changed gallery_id to exhibit_id
        back_populates='exhibit',  # Use back_populates, not backref
        lazy=True
    )

    def __init__(self, name: str = "Exhibit", user_id: int = None, guest_id: str = None):  # Changed default name
        self.name = name
        self.user_id = user_id
        self.guest_id = guest_id

    @classmethod
    def cleanup_guest_exhibits(cls, hours=24):  # Renamed from cleanup_guest_galleries
        """Clean up guest exhibits older than specified hours"""
        from datetime import datetime, timedelta
        cutoff = datetime.utcnow() - timedelta(hours=hours)
        guest_users = User.query.filter(
            User.is_guest == True,
            User.created_at < cutoff
        ).all()
        for user in guest_users:
            cls.query.filter_by(user_id=user.id).delete()
        db.session.commit()

    def add_wall(self, name: str, width: float, height: float, color: str = "White") -> "Wall":
        from .wall import Wall
        wall = Wall(name=name, width=width, height=height, color=color, exhibit_id=self.id)
        db.session.add(wall)
        return wall

    def get_walls(self) -> list["Wall"]:
        return self.walls

    def get_wall_by_name(self, name: str) -> "Wall | None":
        from .wall import Wall
        return Wall.query.filter_by(name=name, exhibit_id=self.id).first()

    def remove_wall(self, wall_id: int) -> bool:
        wall = Wall.query.get(wall_id)
        if wall and wall.exhibit_id == self.id:
            db.session.delete(wall)
            return True
        return False

    def add_unplaced_artwork(self, title: str, medium: str = "", width: float = 0, 
                            height: float = 0, depth: float = 0, price: float = 0,
                            nfs: bool = False, notes: str = "", user_id: Optional[int] = None) -> Artwork:
        artwork = Artwork(
            name=title,
            medium=medium,
            width=width,
            height=height,
            depth=depth,
            price=price,
            nfs=nfs,
            notes=notes,
            exhibit_id=self.id,  # Changed from gallery_id
            user_id=user_id
        )
        db.session.add(artwork)
        return artwork

    def get_unplaced_artworks(self) -> List[Artwork]:
        return self.unplaced_artworks

    def place_artwork(self, artwork_id: int, wall_id: int) -> bool:
        artwork = Artwork.query.get(artwork_id)
        wall = Wall.query.get(wall_id)
        if artwork and wall and artwork.exhibit_id == self.id and wall.exhibit_id == self.id:
            artwork.wall_id = wall_id
            return True
        return False

    def unplace_artwork(self, artwork_id: int) -> bool:
        artwork = Artwork.query.get(artwork_id)
        if artwork and artwork.exhibit_id == self.id:
            artwork.wall_id = None
            return True
        return False

    def export_to_excel(self, filename: str = "exhibit_export.xlsx") -> str:  # Changed from gallery_export        """Export gallery data to Excel file"""
        """Export exhibit data to Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Artworks"

        # Add wall information
        for wall in self.walls:
            ws.append([wall.name, wall.width, wall.height, wall.color])
        ws.append([])  # Empty row before gallery title

        # Gallery title
        ws.append([self.name])
        title_row = ws.max_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
        
        ws.append([])  # Empty row before headers
        
        # Column headers
        headers = ["ID", "Name", "Photo", "Medium", "Width", "Height", "Depth", "Value", "NFS", "Notes"]
        colors = ["ADD8E6", "90EE90", "ADD8E6", "FFFF99", "FFFF99", "FFFF99", "FFFF99", "FA8072", "D8BFD8", "FFFFFF"]
        
        header_row = ws.max_row + 1
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Add placed artworks
        art_counter = 0
        for wall in self.walls:
            for artwork in wall.artworks:
                art_counter += 1
                ws.append([
                    artwork.id or art_counter,
                    artwork.title,
                    "",
                    artwork.medium,
                    artwork.width,
                    artwork.height,
                    artwork.depth,
                    artwork.price,
                    artwork.nfs,
                    artwork.notes
                ])

        # Add unplaced artworks section
        ws.append([])
        ws.append(["Unplaced Artwork"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=10)
        title_cell = ws.cell(row=ws.max_row, column=1)
        title_cell.font = Font(size=12, bold=True)
        title_cell.fill = PatternFill(start_color="F0E68C", end_color="F0E68C", fill_type="solid")

        # Add headers again for unplaced artwork
        header_row = ws.max_row + 1
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Add unplaced artworks
        for artwork in self.unplaced_artworks:
            art_counter += 1
            ws.append([
                artwork.id or art_counter,
                artwork.title,
                "",
                artwork.medium,
                artwork.width,
                artwork.height,
                artwork.depth,
                artwork.price,
                artwork.nfs,
                artwork.notes
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        wb.save(filename)
        return filename

    @classmethod
    def import_from_excel(cls, filename: str = "exhibit_export.xlsx") -> 'Exhibit':  # Changed return type
        """Import exhibit data from Excel file"""
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found: {filename}")

        ws = wb["Artworks"]
        exhibit = cls()  # Changed from gallery
        exhibit.name = ws.cell(row=row_idx, column=1).value or "Imported Exhibit"  # Changed from "Imported Gallery"

        # Read wall information
        row_idx = 1
        while ws.cell(row=row_idx, column=1).value:
            wall_name = ws.cell(row=row_idx, column=1).value
            wall_width = float(ws.cell(row=row_idx, column=2).value)
            wall_height = float(ws.cell(row=row_idx, column=3).value)
            wall_color = ws.cell(row=row_idx, column=4).value or "White"
            exhibit.add_wall(name=wall_name, width=wall_width, height=wall_height, color=wall_color)
            row_idx += 1
        
        # Skip empty rows and get gallery name
        row_idx += 1
        exhibit.name = ws.cell(row=row_idx, column=1).value or "Imported Gallery"
        row_idx += 2  # Skip empty row and headers

        reading_unplaced = False
        current_wall = exhibit.walls[-1] if exhibit.walls else None

        for row in ws.iter_rows(min_row=row_idx, values_only=True):
            if not any(row):
                continue
            if row[0] == "Unplaced Artwork":
                reading_unplaced = True
                continue
            if row[0] == "ID":  # Skip header rows
                continue

            # Create artwork from row data
            art_id, title, _, medium, width, height, depth, price, nfs, notes = row
            artwork_data = {
                'title': title or "",
                'medium': medium or "",
                'width': float(width or 0),
                'height': float(height or 0),
                'depth': float(depth or 0),
                'price': float(price or 0),
                'nfs': bool(nfs),
                'notes': notes or "",
                'exhibit_id': exhibit.id
            }

            if reading_unplaced:
                exhibit.add_unplaced_artwork(**artwork_data)
            elif current_wall:
                artwork = Artwork(**artwork_data, wall_id=current_wall.id)
                db.session.add(artwork)

        return exhibit

    def __repr__(self):
        return f"<Exhibit {self.name} ({len(self.walls)} walls, {len(self.unplaced_artworks)} unplaced artworks)>"