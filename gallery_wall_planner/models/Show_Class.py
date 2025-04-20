import openpyxl
from openpyxl.styles import Font, PatternFill
from config_create_read import read_config

class Show:
    # Class level list of all galleries in the show
    # This may not be the correct way to name things but we can always refactor if needed
    _all_galleries = []
    # Constructor
    def __init__(self, name="Roberta's Gallery"):
        self.name = name
        self.galleries = []

    @classmethod
    def add_gallery(cls, gallery):
        cls._all_galleries.append(gallery)
        
    @classmethod
    def get_galleries(cls):
        return cls._all_galleries
        
    @classmethod
    def remove_gallery(cls, gallery):
        cls._all_galleries.remove(gallery)

    @classmethod
    def get_gallery_by_name(cls, name):
        for gallery in cls._all_galleries:
            if gallery.name == name:
                return gallery
        return None

    def export_show(self, filename="show_export.xlsx"):
        """
        Export Show Function
        Inputs: Self and a name for the export
        Process: Automatically writes all artwork, wall data and gallery data to an excel file
        Outputs: An excel file with each gallery on a seperate sheet, with all relevant wall and artwork data
        """
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)  # Remove default sheet
        # Getting data from config file for headers and colors
        config_values = read_config()
        
        # Creating seperate sheets for each gallery
        for gallery in self.galleries:
            ws = wb.create_sheet(title=gallery.name[:31])  # Excel limits sheet names to 31 chars
            
            # Gallery title
            ws.append([gallery.name])
            title_row = ws.max_row
            ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
            title_cell = ws.cell(row=title_row, column=1)
            title_cell.font = Font(size=14, bold=True)
            title_cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
            ws.append([])

            for wall in gallery.walls:
                ws.append([wall.name, wall.width, wall.height, getattr(wall, "color", "")])
            ws.append([])
            # Replaced with a call to the config file to get headers and colors
            headers = config_values['headers']
            colors = config_values['header_colors']
            # Adding headers
            header_row = ws.max_row + 1
            for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            # Getting all walls and adding any artwork associated with them for the excel
            for wall in gallery.walls:
                for artwork in wall.artwork:
                    ws.append([
                        getattr(artwork, "id", ""),
                        artwork.title,
                        "",
                        artwork.medium,
                        artwork.width,
                        artwork.height,
                        artwork.depth,
                        artwork.price,
                        artwork.nfs,
                        getattr(artwork, "notes", "")
                    ])
            ws.append([])
        
        wb.save(filename)
        print(f"Show exported to {filename}")

    @classmethod
    def import_show(cls, filename="show_export.xlsx"):
        """
        Import show function
        Input: The object we're storing to and the name of the file to import from
        Process: Gets all artwork, wall and gallery data from an excel sheet (if in correct format)
        Output: A show object with all gallery, wall and artwork data
        """
        wb = openpyxl.load_workbook(filename)
        show = cls(name=wb.sheetnames[0])
        # Get gallery from sheet names
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            gallery = Gallery(name=sheet_name)
            show.galleries.append(gallery)
            row_idx = 2
            # Get all the walls from the sheet
            while row_idx <= ws.max_row:
                wall_name = ws.cell(row=row_idx, column=1).value
                if wall_name:
                    wall_width = ws.cell(row=row_idx, column=2).value
                    wall_height = ws.cell(row=row_idx, column=3).value
                    wall_color = ws.cell(row=row_idx, column=4).value
                    gallery.walls.append(Wall(name=wall_name, width=wall_width, height=wall_height, color=wall_color))
                row_idx += 1
            
            row_idx += 1  # Move past the empty row after walls
            # Get all of the artwork data from the walls
            for row in ws.iter_rows(min_row=row_idx, values_only=True):
                if not any(row):
                    continue  # Skip empty rows
                
                if row[0] == "ID":
                    continue  # Skip header row
                
                art_id, title, _, medium, width, height, depth, price, nfs, notes = row
                artwork = Artwork(
                    title=title or "",
                    medium=medium or "",
                    width=width or 0,
                    height=height or 0,
                    depth=depth or 0,
                    price=price or 0,
                    nfs=bool(nfs)
                )
                setattr(artwork, "id", art_id)
                setattr(artwork, "notes", notes or "")
                if gallery.walls:
                    gallery.walls[-1].artwork.append(artwork)
            
        print(f"Show '{show.name}' imported successfully with {len(show.galleries)} galleries.")
        return show
