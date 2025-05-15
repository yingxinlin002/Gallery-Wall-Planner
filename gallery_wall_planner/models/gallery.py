import openpyxl
from openpyxl.styles import Font, PatternFill
from typing import List, Optional, Union, Dict, Any
from gallery_wall_planner.models.wall import Wall
from gallery_wall_planner.models.artwork import Artwork

class Gallery:
    """Class representing a gallery with multiple walls and artworks."""
    
    def __init__(self, name: str = "Gallery"):
        self.name = name  # Exhibit title
        self.walls = []  # List of Wall objects for this specific gallery
        self.walls_dict: Dict[str, Wall] = {}
        self.current_wall: Optional[Wall] = None
        self.unplaced_artwork: List[Artwork] = []

    # Class methods to manage all walls
    def add_wall(self, wall: Wall) -> None:
        self.walls.append(wall)
        self.walls_dict[wall.id] = wall
        self.current_wall = wall
        
    def get_walls(self) -> List[Wall]:
        return self.walls

    def get_walls_dict(self) -> Dict[str, Wall]:
        return self.walls_dict
        
    def remove_wall(self, wall: Wall) -> None:
        if self.current_wall == wall:
            self.current_wall = None
        self.walls.remove(wall)
        self.walls_dict.pop(wall.id)

    def update_wall(self, wall_id : str, wall : Wall) -> bool:

        if wall_id in self.walls_dict:
            old_wall = self.walls_dict.pop(wall_id)
            self.walls.remove(old_wall)
            self.walls.append(wall)
            self.walls_dict[wall.id] = wall
            self.current_wall = wall
            return True
        return False

    def get_wall_by_name(self, name: str) -> Optional[Wall]:
        """Find a wall by its name"""
        for wall in self.walls:
            if wall.name == name:
                return wall
        return None
        
    def add_unplaced_artwork(self, artwork: Artwork):
        self.unplaced_artwork.append(artwork)
        
    def get_art_by_name(self, name: str):
        for art in self.unplaced_artwork:
            if art.name == name:
                return art
                
    def remove_artwork(self, artwork: Artwork):
        self.unplaced_artwork.remove(artwork)
        
    def place_art(self, art_name: str, wall_name: str):
        """
        Args: self, The name of the art to place, the name of the wall to place art upon
        Returns: None
        """
        art = self.get_art_by_name(art_name)
        wall = self.get_wall_by_name(wall_name)
        wall.add_artwork(art)
        self.remove_artwork(art)

    def unplace_art(self, art: Artwork, wall_name: str):
        """
        Args: self, The artwork object to un-place, the name of the wall to remove the art from
        Returns: None
        """
        wall = self.get_wall_by_name(wall_name)
        self.add_unplaced_artwork(art)
        wall.remove_artwork(art)
        
    def export_gallery(self, filename="gallery_export.xlsx"):
        """
        Export the gallery to an Excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Artworks"

        # Add wall information first
        for wall in self.walls:
            ws.append([wall.name, wall.width, wall.height, getattr(wall, "color", "")])
        ws.append([])  # Empty row before exhibit title

        # Exhibit title
        ws.append([self.name])
        title_row = ws.max_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
        
        ws.append([])  # Empty row before headers
        
        headers = ["ID", "Name", "Photo", "Medium", "Width", "Height", "Depth", "Value", "NFS", "Notes"]
        colors = ["ADD8E6", "90EE90", "ADD8E6", "FFFF99", "FFFF99", "FFFF99", "FFFF99", "FA8072", "D8BFD8", "FFFFFF"]
        #Loop to apply header styles
        header_row = ws.max_row + 1
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        #Counter for artwork ID's
        art_counter = 0
        #Loop thru walls
        for wall in self.walls:
            #Loop thru artwork inside the walls
            for artwork in wall.artwork:
                art_counter += 1
                ws.append([
                    getattr(artwork, "id", art_counter),
                    artwork.title,
                    "",
                    artwork.medium,
                    artwork.width,
                    artwork.height,
                    artwork.depth,
                    artwork.price,
                    artwork.nfs,
                    artwork.notes
                ])
        #Apply artwork data
        for col in ws.iter_cols(min_row=header_row, max_row=ws.max_row):
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Gallery exported to {filename}")

        # Unplaced Artworks marker
        ws.append([])
        ws.append(["Unplaced Artwork"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=10)
        title_cell = ws.cell(row=ws.max_row, column=1)
        title_cell.font = Font(size=12, bold=True)
        title_cell.fill = PatternFill(start_color="F0E68C", end_color="F0E68C", fill_type="solid")

        # Add headers again for unplaced artwork
        header_row = ws.max_row + 1
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Append unplaced artworks
        for artwork in self.unplaced_artwork:
            art_counter += 1
            ws.append([
                getattr(artwork, "id", art_counter),
                artwork.title,
                "",
                artwork.medium,
                artwork.width,
                artwork.height,
                artwork.depth,
                artwork.price,
                artwork.nfs,
                artwork.notes
            ])
        
    @classmethod
    def import_gallery(cls, filename="gallery_export.xlsx"):
        """Import a gallery from an Excel file."""
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            raise FileNotFoundError(f"GallerExcelNotFoundError: gallery excel file with specific name or path not found: {filename}")
        ws = wb["Artworks"]
        #Initialize row and gallery
        row_idx = 1
        gallery = cls()

        # Read wall information
        while ws.cell(row=row_idx, column=1).value:
            wall_name = ws.cell(row=row_idx, column=1).value
            wall_width = ws.cell(row=row_idx, column=2).value
            wall_height = ws.cell(row=row_idx, column=3).value
            wall_color = ws.cell(row=row_idx, column=4).value
            imported_wall = Wall(name=wall_name, width=wall_width, height=wall_height, color=wall_color)
            gallery.walls.append(imported_wall)
            row_idx += 1
        
        row_idx += 1  # Skip empty row before exhibit title
        gallery.name = ws.cell(row=row_idx, column=1).value or "Imported Exhibit" #Dynamically get gallery name
        
        row_idx += 2  # Skip empty row and headers
        
        reading_unplaced = False
        current_wall = gallery.walls[-1] if gallery.walls else None

        for row in ws.iter_rows(min_row=row_idx, values_only=True):
            if not any(row):
                continue
            if row[0] == "Unplaced Artwork":
                reading_unplaced = True
                continue
            if row[0] == "ID":  # Header row (reappears before unplaced)
                continue

            # Unpack artwork row
            art_id, title, _, medium, width, height, depth, price, nfs, notes = row
            artwork = Artwork(
                title=title or "",
                medium=medium or "",
                width=width or 0,
                height=height or 0,
                depth=depth or 0,
                price=price or 0,
                nfs=bool(nfs),
                notes=notes or ""
            )
            setattr(artwork, "id", art_id)

            if reading_unplaced:
                gallery.unplaced_artwork.append(artwork)
            elif current_wall:
                current_wall.artwork.append(artwork)

        print(f"Gallery '{gallery.name}' imported successfully with {sum(len(w.artwork) for w in gallery.walls)} artworks.")
        print(f"Unplaced artworks: {len(gallery.unplaced_artwork)}")
        return gallery
