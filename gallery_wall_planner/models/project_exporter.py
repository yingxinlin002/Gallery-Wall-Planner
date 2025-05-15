# project_exporter.py

import json
import os
import ast
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from gallery_wall_planner.models.wall import Wall
from gallery_wall_planner.models.artwork import Artwork
from gallery_wall_planner.models.permanent_object import PermanentObject
from gallery_wall_planner.models.structures import Position
from gallery_wall_planner.models.gallery import Gallery  # assuming this exists
from .wall_line import SingleLine
from gallery_wall_planner.models.gallery import Gallery
from gallery_wall_planner.models.wall_line import Orientation
from typing import Dict


def export_project(filepath, wall, permanent_objects, layout):
    """
    Save the current project data (wall, permanent objects, layout) into a JSON file.
    
    Args:
        filepath (str): Path to save the JSON file.
        wall (Wall): The Wall object.
        permanent_objects (list): List of permanent objects (dicts).
        layout (dict): Dictionary of layout positions keyed by object names.
    """
    project_data = {
        "wall": wall.export() if hasattr(wall, 'export') else wall.__dict__,  # fallback if export() is not defined
        "permanent_objects": permanent_objects,
        "layout": layout
    }
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    with open(filepath, "w") as f:
        json.dump(project_data, f, indent=4)
    print(f"[INFO] Project saved to {filepath}")

def import_project(filepath):
    """
    Load a project from a JSON file.
    
    Args:
        filepath (str): Path to the JSON file.
    
    Returns:
        dict: Dictionary containing wall, permanent_objects, and layout.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Project file not found: {filepath}")

    with open(filepath, "r") as f:
        project_data = json.load(f)

    print(f"[INFO] Project loaded from {filepath}")
    return project_data
    
def _project_to_excel(filepath, wall=None, artworks=None, permanent_objects=None):
    """
    Export project data to Excel. Any of wall, artworks, or permanent_objects may be None.

    Args:
        filepath (str): Destination Excel path
        wall (Wall): Wall object
        artworks (list of Artwork): Artwork objects
        permanent_objects (list of PermanentObject): Permanent objects
    """
    def to_visible_wall_data(wall):
        """Convert wall data to a visible format for Excel export."""
        return {
            "Wall Name": wall.name,
            "Width (in)": wall.width,
            "Height (in)": wall.height,
        }

    def to_visible_artwork_data(art):
        """Convert artwork data to a visible format for Excel export."""
        return {
            "Name": art.name,
            "Width": art.width,
            "Height": art.height,
            "Hanging Point": art.hanging_point,
            "Medium": art.medium,
            "Depth": art.depth,
            "Image_Path": art.image_path,
            "NFS (Y/N)": "Y" if art.nfs else "N"
        }

    def to_internal_data(obj):
        """Convert object data to a safe format for Excel export."""
        try:
            d = obj.export()
        except AttributeError:
            d = obj.__dict__
        safe = {}
        for k, v in d.items():
            if isinstance(v, Position):
                safe[k] = {"x": v.x, "y": v.y}
            else:
                safe[k] = v
        return str(safe)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        if wall:
            wall_data = pd.DataFrame([to_visible_wall_data(wall)])
            wall_data["_internal"] = to_internal_data(wall)
            wall_data.to_excel(writer, sheet_name="Wall", index=False)

        if artwork:
            artworks_data = pd.DataFrame([to_visible_artwork_data(a) for a in artwork])
            artworks_data["_internal"] = [to_internal_data(a) for a in artwork]
            artworks_data.to_excel(writer, sheet_name="Artworks", index=False)

        if permanent_objects:
            perms_data = pd.DataFrame([p.__dict__ for p in permanent_objects])
            perms_data["_internal"] = [to_internal_data(p) for p in permanent_objects]
            perms_data.to_excel(writer, sheet_name="Permanents", index=False)

    print(f"[INFO] Project exported to Excel at {filepath}")

    def to_internal_data(obj):
        """Convert object data to a safe format for Excel export."""
        safe = {}
        for k, v in obj.__dict__.items():
            if isinstance(v, Position):
                safe[k] = {"x": v.x, "y": v.y}
            else:
                safe[k] = v
        return str(safe)
    # Store data in pandas dataframe
    wall_data = pd.DataFrame([to_visible_wall_data(wall)])
    wall_data["_internal"] = to_internal_data(wall)

    artworks_data = pd.DataFrame([to_visible_artwork_data(a) for a in artwork])
    artworks_data["_internal"] = [to_internal_data(a) for a in artwork]

    perms_data = pd.DataFrame([p.__dict__ for p in permanent_objects])
    perms_data["_internal"] = [to_internal_data(p) for p in permanent_objects]
    # Write pandas dataframe to excel using openpyxl
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        wall_data.to_excel(writer, sheet_name="Wall", index=False)
        artworks_data.to_excel(writer, sheet_name="Artworks", index=False)
        perms_data.to_excel(writer, sheet_name="Permanents", index=False)

    print(f"[INFO] Project exported to Excel at {filepath}")

    def import_gallery_from_excel(filepath):
        """Import a gallery from an Excel file."""
        print(f"[INFO] Importing gallery from {filepath}")
        xls = pd.ExcelFile(filepath)
        gallery = Gallery(name="Imported Gallery")
    
        walls = {}
    
        # Load wall metadata first
        if "Walls" in xls.sheet_names:
            df_walls = xls.parse("Walls")
            for _, row in df_walls.iterrows():
                name = row["name"]
                width = float(row["width"])
                height = float(row["height"])
                color = row.get("color", "#FFFFFF")
                walls[name] = Wall(name=name, width=width, height=height, color=color)
        else:
            print("[WARN] No 'Walls' sheet found. Walls will use default dimensions.")
    
        # Now process each wall's data
        for sheet in xls.sheet_names:
            print(f"[INFO] Reading sheet: {sheet}")
            df = xls.parse(sheet)
            if df.empty:
                print(f"[WARN] Sheet '{sheet}' is empty, skipping.")
                continue
    
            if " - Art" in sheet:
                wall_name = sheet.replace(" - Art", "")
                if df.columns.tolist() == ["info"] and "No artworks" in df["info"].values:
                    continue
                if wall_name in walls:
                    walls[wall_name].artwork = [Artwork.from_dict(row) for _, row in df.iterrows()]
                    print(f"[OK] Loaded artworks for wall '{wall_name}'.")
                else:
                    print(f"[ERROR] Wall '{wall_name}' not found in Walls sheet.")
    
            elif " - Lines" in sheet:
                wall_name = sheet.replace(" - Lines", "")
                if df.columns.tolist() == ["info"] and "No wall lines" in df["info"].values:
                    continue
                if wall_name in walls:
                    walls[wall_name].wall_lines = [SingleLine.from_dict(row) for _, row in df.iterrows()]
                    print(f"[OK] Loaded wall lines for wall '{wall_name}'.")
                else:
                    print(f"[ERROR] Wall '{wall_name}' not found in Walls sheet.")
    
            elif " - Perm" in sheet:
                wall_name = sheet.replace(" - Perm", "")
                if df.columns.tolist() == ["info"] and "No permanent objects" in df["info"].values:
                    continue
                if wall_name in walls:
                    walls[wall_name].permanent_objects = [PermanentObject.from_dict(row) for _, row in df.iterrows()]
                    print(f"[OK] Loaded permanent objects for wall '{wall_name}'.")
                else:
                    print(f"[ERROR] Wall '{wall_name}' not found in Walls sheet.")
    
        gallery.walls = list(walls.values())
        print(f"[DONE] Imported gallery with {len(gallery.walls)} walls.")
        return gallery



def export_gallery_to_excel(filepath: str, gallery: Gallery):
    """Export the gallery to an Excel file."""
    print(f"[INFO] Exporting gallery to {filepath}")
    writer = pd.ExcelWriter(filepath, engine='xlsxwriter')

    # Sheet: ExportInfo
    pd.DataFrame([{"Export status": "Initialized"}]).to_excel(writer, index=False, sheet_name="ExportInfo")

    # NEW Sheet: Walls
    wall_rows = []
    for wall in gallery.walls:
        wall_rows.append({
            "name": wall.name,
            "width": wall.width,
            "height": wall.height,
            "color": wall.color
        })
    pd.DataFrame(wall_rows).to_excel(writer, index=False, sheet_name="Walls")

    # Export each wall's elements
    for wall in gallery.walls:
        # Artwork
        if wall.artwork:
            def artwork_to_export_dict(artwork):
                d = artwork.to_dict()
                d["x"] = getattr(artwork.position, "x", None)
                d["y"] = getattr(artwork.position, "y", None)
                return d

            pd.DataFrame([artwork_to_export_dict(a) for a in wall.artwork]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Art"
            )
        else:
            pd.DataFrame([{"info": "No artworks"}]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Art"
            )

        # Wall Lines
        if wall.wall_lines:
            for line in wall.wall_lines:
                print(line.to_dict())
            pd.DataFrame([l.to_dict() for l in wall.wall_lines]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Lines"
            )
        else:
            pd.DataFrame([{"info": "No wall lines"}]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Lines"
            )

        # Permanent Objects
        if wall.permanent_objects:
            pd.DataFrame([p.to_dict() for p in wall.permanent_objects]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Perm"
            )
        else:
            pd.DataFrame([{"info": "No permanent objects"}]).to_excel(
                writer, index=False, sheet_name=f"{wall.name} - Perm"
            )

    writer.close()
    print("[DONE] Finished exporting gallery.")

def import_gallery_from_excel(filepath: str) -> Gallery:
    """Import a gallery from an Excel file."""
    print(f"[INFO] Importing gallery from {filepath}")
    wb: Workbook = openpyxl.load_workbook(filepath, data_only=True)
    gallery = Gallery("Imported Gallery")

    wall_map: Dict[str, Wall] = {}  # name -> Wall object

    # Step 1: Load walls from "Walls" sheet
    if "Walls" in wb.sheetnames:
        sheet: Worksheet = wb["Walls"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, width, height, color = row[:4]
            if name:
                wall = Wall(name=name, width=width, height=height, color=color)
                gallery.add_wall(wall)
                wall_map[name] = wall
                print(f"[INFO] Loaded wall: {name}")
    else:
        print("[ERROR] 'Walls' sheet not found.")
        return gallery

    # Step 2: Loop through each sheet and map to walls
    for sheet_name in wb.sheetnames:
        if sheet_name in ["Walls", "ExportInfo"]:
            continue

        try:
            if " - Art" in sheet_name:
                wall_name = sheet_name.replace(" - Art", "")
                wall = wall_map.get(wall_name)
                if not wall:
                    print(f"[WARN] Art sheet found for unknown wall '{wall_name}' — skipping")
                    continue
                sheet: Worksheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, width, height, hanging_point, medium, depth, image_path, nfs, notes, price, x, y = row[:12]
                    if name:
                        artwork = Artwork(name=name, width=width, height=height,
                                          hanging_point=hanging_point,
                                          medium=medium or "",
                                          depth=depth or 0.0,
                                          image_path=image_path if isinstance(image_path, str) else None,
                                          nfs=(str(nfs).strip().upper() == "Y"))

                        if isinstance(x, (int, float)) and isinstance(y, (int, float)):
                            artwork.position = Position(x, y)

                        wall.add_artwork(artwork)
                print(f"[INFO] Imported artwork for wall '{wall_name}'")

            elif " - Lines" in sheet_name:
                wall_name = sheet_name.replace(" - Lines", "")
                wall = wall_map.get(wall_name)
                if not wall:
                    print(f"[WARN] Line sheet found for unknown wall '{wall_name}' — skipping")
                    continue
                sheet: Worksheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    x_cord, y_cord, length, angle, snap_to, moveable, orientation, alignment, distance = row[:9]
                    line_orientation: Orientation = Orientation.HORIZONTAL if orientation == "horizontal" else Orientation.VERTICAL
                    if x_cord is not None and y_cord is not None:
                        wall_line = SingleLine(distance=distance, orientation=line_orientation)
                        wall.add_wall_line(wall_line)
                        print(f"[DEBUG] Adding wall line to wall: {wall_line}")
                print(f"[INFO] Imported wall lines for wall '{wall_name}'")

            elif " - Perm" in sheet_name:
                wall_name = sheet_name.replace(" - Perm", "")
                wall = wall_map.get(wall_name)
                if not wall:
                    print(f"[WARN] Perm sheet found for unknown wall '{wall_name}' — skipping")
                    continue
                sheet: Worksheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    values = list(row)
                    if len(values) < 5:
                        print(f"[WARN] Skipping incomplete permanent object row: {values}")
                        continue
                
                    name = values[0]
                    
                    # Try parsing (x, y) from a single string
                    x = y = None
                    if isinstance(values[2], str) and "(" in values[2] and "," in values[2]:
                        try:
                            coord_str = values[2].strip("() ")
                            x_str, y_str = coord_str.split(",")
                            x = float(x_str.strip())
                            y = float(y_str.strip())
                        except Exception as e:
                            print(f"[ERROR] Could not parse position from '{values[2]}': {e}")
                    else:
                        x = values[1]
                        y = values[2]
                    
                    width = float(values[3]) if values[3] is not None else None
                    height = float(values[4]) if values[4] is not None else None
                    image_path = values[5] if len(values) > 5 else None
                
                    if name and x is not None and y is not None:
                        safe_image_path = image_path if isinstance(image_path, str) and image_path.strip() else None
                        perm = PermanentObject(name=name, width=width, height=height, image_path=safe_image_path)
                        perm.position = Position(x, y)
                        wall.add_permanent_object(perm)

                        print(f"[DEBUG] Adding permanent object to wall: {perm}")
                    
                print(f"[INFO] Imported permanent objects for wall '{wall_name}'")
                print(f"[DEBUG] Total permanent objects: {len(wall.permanent_objects)}")


        except Exception as e:
            print(f"[ERROR] Failed to load sheet '{sheet_name}': {e}")

    print(f"[DONE] Finished importing gallery with {len(gallery.walls)} walls.")
    return gallery

    gallery.walls = list(walls.values())
    print(f"[DONE] Finished importing gallery with {len(gallery.walls)} walls.")
    return gallery

# All available methods    
__all__ = [
    "export_project",
    "import_project",
    "export_project_to_excel",
    "import_project_from_excel",
    "export_gallery_to_excel",
    "import_gallery_from_excel"
]
