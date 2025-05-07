import json
import re
import os
import openpyxl
from types import SimpleNamespace
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from gallery_wall_planner.models.wall_object import WallObject
from gallery_wall_planner.models.structures import get_id
from typing import override

class Artwork(WallObject):
    """
    Represents an artwork that can be placed on a wall
    """
    def __init__(self, name: str = "", medium: str = "", height: float = 0.0, width: float = 0.0, 
                 depth: float = 0.0, hanging_point: float = 0.0, price: float = 0.0, 
                 nfs: bool = False, image_path: str = "", notes: str = ""):
        # Initialize the parent class with common attributes
        super().__init__(name, width, height, image_path)
        
        # Initialize artwork-specific private attributes
        self._medium = medium
        self._depth = depth
        self._hanging_point = hanging_point
        self._price = price
        self._nfs = nfs
        self._notes = notes

        # Ensure height is always a float
        self.height = height
        
    # name property is inherited from WallObject
    
    @override
    def _get_id(self):
        return get_id("artwork"+self.name+f"width{self.width},height{self.height},hanging_point{self.hanging_point}")
    
    @property
    def medium(self) -> str:
        """Get the artwork's medium"""
        return self._medium
    
    @medium.setter
    def medium(self, value: str):
        """Set the artwork's medium with validation"""
        if not isinstance(value, str):
            raise ValueError("Medium must be a string")
        self._medium = value
    
    @property
    def height(self) -> float:
        """Get the artwork's height"""
        return self._height if hasattr(self, '_height') else 0.0
    
    @height.setter
    def height(self, value: float):
        """Set the artwork's height with validation"""
        if not isinstance(value, (int,float)):
            raise ValueError("height must be a number")
        if value <= 0:
            raise ValueError("Height must be a positive value")
        self._height = float(value) if value is not None else 0.0
    
    @property
    def depth(self) -> float:
        """Get the artwork's depth"""
        return self._depth
    
    @depth.setter
    def depth(self, value: float):
        """Set the artwork's depth with validation"""
        if not isinstance(value, (int, float)):
            raise ValueError("Depth must be a number")
        if value < 0:
            raise ValueError("Depth cannot be negative")
        self._depth = float(value)
    
    @property
    def hanging_point(self) -> float:
        """Get the artwork's hanging point"""
        return self._hanging_point
    
    @hanging_point.setter
    def hanging_point(self, value: float):
        """Set the artwork's hanging point with validation"""
        if not isinstance(value, (int, float)):
            raise ValueError("Hanging point must be a number")
        if value < 0:
            raise ValueError("Hanging point cannot be negative")
        self._hanging_point = float(value)
    
    @property
    def price(self) -> float:
        """Get the artwork's price"""
        return self._price
    
    @price.setter
    def price(self, value: float):
        """Set the artwork's price with validation"""
        if not isinstance(value, (int, float)):
            raise ValueError("Price must be a number")
        if value < 0:
            raise ValueError("Price cannot be negative")
        self._price = float(value)
    
    @property
    def nfs(self) -> bool:
        """Get the artwork's not-for-sale status"""
        return self._nfs
    
    @nfs.setter
    def nfs(self, value: bool):
        """Set the artwork's not-for-sale status with validation"""
        if not isinstance(value, bool):
            raise ValueError("NFS must be a boolean")
        self._nfs = value
    
    # image_path property is inherited from WallObject
    
    @property
    def notes(self) -> str:
        """Get the artwork's notes"""
        return self._notes
    
    @notes.setter
    def notes(self, value: str):
        """Set the artwork's notes with validation"""
        if not isinstance(value, str):
            raise ValueError("Notes must be a string")
        self._notes = value

    def __str__(self):
        #return all info about the artwork
        return f"Name: {self.name}\nMedium: {self.medium}\nHeight: {self.height}\nWidth: {self.width}\nDepth: {self.depth}\nHanging Point: {self.hanging_point}\nPrice: {self.price}\nNFS: {self.nfs}\nImage Path: {self.image_path}\nNotes: {self.notes}"



    def export_to_excel(self, filename="artwork_export.xlsx"):
        """
        Inputs: self, file name 
        Process: Exports all artwork data in order to write to excel sheet, including images if possible
        Outputs: An excel sheet containing a single artwork, I can try to fix this later to do multiple artworks if needed
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Artworks"
        # Making colors and headers
        headers = ["Name", "Medium", "Height", "Width", "Depth", "Hanging Point", "Value", "NFS", "Image", "Notes"]
        colors = ["FFC7CE", "FFF2CC", "FFC7CE", "FFC7CE", "FFF2CC", "FFC7CE", "FFF2CC", "FFF2CC", "FFF2CC", "FFF2CC"]
        # Iterate to make headers
        for col_num, (header, color) in enumerate(zip(headers, colors), start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Append artwork info
        ws.append([self.name, self.medium, self.height, self.width, self.depth, self.hanging_point, self.price, self.nfs, self.image_path, self.notes])
        # This should theoretically embed the image if it exists, otherwise adds the image path instead
        if self.image_path and os.path.exists(self.image_path):
            img = Image(self.image_path)
            img.width, img.height = 100, 100  # Resize for visibility
            ws.add_image(img, "I2")
        
        wb.save(filename)
        print(f"Artwork exported to {filename}")
    
    @classmethod
    def import_from_excel(cls, filename="artwork_export.xlsx", sheetname = "Artworks"):
        """
        Syntax: This needs to be used via the following syntax Artwork.import_from_excel(filename, sheetname)
        Input: class, file name (optional), sheet name (optional) 
        Process: Opens specified excel and specified sheet, gets all artwork data and creates artwork objects, then adds to a list
        Outputs: A list containing artwork objects gotten from the excel sheet
        """
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            raise FileNotFoundError(f"ArtworkExcelNotFoundError: Excel file of the artwork with specific name or path couldn't be found: {filename}")
            
        ws = wb[sheetname] # This can cause errors if the sheet name does not match the actual sheet
        # List of artworks
        artworks = []
        # Get artwork data
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, medium, height, width, depth, hanging_point, price, nfs, image_path, notes = row
            # Gets the image path to place into the image section of artwork 
            if image_path and os.path.exists(image_path):
                img_dir = "imported_images"
                os.makedirs(img_dir, exist_ok=True)
                img_save_path = os.path.join(img_dir, os.path.basename(image_path))
                
                # Save the image from Excel to a file
                pil_img = PILImage.open(image_path)
                pil_img.save(img_save_path)
                image_path = img_save_path
            
            artwork = cls(name, medium, height, width, depth, hanging_point, price, nfs, image_path, notes)
            artworks.append(artwork)
        
        print(f"Imported {len(artworks)} artworks from {filename}")
        return artworks    

    def to_dict(self):
        # Helper for exporter.py
        return {
            "name": self.name,
            "width": self.width,
            "height": self.height,
            "hanging_point": self.hanging_point,
            "medium": self.medium,
            "depth": self.depth,
            "image_path": self.image_path,
            "nfs": self.nfs,
            "notes": self.notes,
            "price": self.price
        }
    
    @staticmethod
    def from_dict(data):
        # Helper for importing for exporter.py
        return Artwork(
            title=data.get("name", ""),
            width=data.get("width", 0),
            height=data.get("height", 0),
            hanging_point=data.get("hanging point", 0),
            medium=data.get("medium", ""),
            depth=data.get("depth", 0),
            photo=data.get("photo", ""),
            nfs=(data.get("NFS (Y/N)", "").strip().upper() == "Y")
        )

    def export_artwork(self, directory: str = "") -> str:
        """ Legacy method to Export artwork to a JSON file
        
        Args:
            directory (str, optional): Directory to save the file. Defaults to current directory.
        """
        # Preprocess name to make it a valid filename
        # Replace spaces and special characters with underscores
        safe_name = re.sub(r'[^\w\-\.]', '_', self.name)
        # Remove multiple consecutive underscores
        safe_name = re.sub(r'_+', '_', safe_name)
        # Remove leading/trailing underscores
        safe_name = safe_name.strip('_')
        # Default to 'artwork' if name is empty after processing
        if not safe_name:
            safe_name = "artwork"
            
        # Create the full file path
        file_path = os.path.join(directory, f"{safe_name}_artwork_export.json")
        
        # Export to JSON
        with open(file_path, "w") as f:
            f.write(json.dumps(self.__dict__))
            
        return file_path

def import_artwork(art_name, file_name):
    # Legacy method, not going to remove yet as it might cause issues, but we should eventually remove this
    try:
        with open(file_name, "rb") as f:
            data = f.read()
    except FileNotFoundError:
        raise FileNotFoundError(f"ArtworkJsonNotFoundError: Json file of artwork with specific name or path not found: {file_name}")
    art_name = json.loads(data, object_hook=lambda d: SimpleNamespace(**d))

